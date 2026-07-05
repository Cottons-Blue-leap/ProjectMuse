#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
analytics_xlsx.py — Project Muse 채널 분석 Excel 리포트 생성기

Analytics/ 폴더의 CSV 시계열(snapshots/traffic/geo/search)을 읽어
한 권의 분석 워크북(channel_analysis.xlsx)으로 렌더한다.

- 입력: youtube_analytics.py report 가 쌓는 CSV (라이브 데이터의 사본).
- 출력: Analytics/channel_analysis.xlsx (매 실행 재생성 = 라이브와 drift 없음).
- 숫자는 검증된 CSV에서 그대로 옮겨 적음 (LibreOffice 미설치 환경 → 수식 대신 값,
  열기 즉시 정확/차트 즉시 렌더). Studio 수동 보충 칸만 노란 빈 칸으로 남김.

usage:  python analytics_xlsx.py            # Analytics/ 기본 경로 (스크립트 옆)
        python analytics_xlsx.py <dir>      # 다른 폴더
또는 youtube_analytics.py report 가 자동 호출.
"""

import sys
import csv
import datetime as dt
from pathlib import Path

# Windows cp949 한글 깨짐 방어 (s355 광역 audit)
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

# ---- 스타일 (Arial · 미쿠 청록 헤더) ----
FONT = "Arial"
TEAL = "1F6F6B"
TITLE_F = Font(name=FONT, bold=True, size=16, color=TEAL)
SUB_F = Font(name=FONT, italic=True, size=9, color="666666")
H_F = Font(name=FONT, bold=True, size=10, color="FFFFFF")
SEC_F = Font(name=FONT, bold=True, size=12, color=TEAL)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, bold=True, size=10)
NOTE_F = Font(name=FONT, size=9, color="555555")
H_FILL = PatternFill("solid", fgColor=TEAL)
ALT_FILL = PatternFill("solid", fgColor="EAF2F1")
YELLOW = PatternFill("solid", fgColor="FFFF00")
KPI_FILL = PatternFill("solid", fgColor="F4F0E4")  # cream
_thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
RIGHT = Alignment(horizontal="right")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

INT_FMT = "#,##0"
PCT_FMT = '0.0"%"'
SEC_FMT = '0"s"'


def _read(path):
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def _num(v):
    if v in (None, ""):
        return None
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 2)
    except (TypeError, ValueError):
        return v


def _set(ws, r, c, val, font=BODY, fmt=None, fill=None, align=None, border=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = BORDER
    return cell


def _table(ws, top, left, headers, rows, fmts):
    """헤더 + 데이터 표. fmts = [(number_format or None, Font/None), ...] per column."""
    for j, h in enumerate(headers):
        _set(ws, top, left + j, h, font=H_F, fill=H_FILL,
             align=CENTER if j else LEFT)
    for i, row in enumerate(rows):
        rr = top + 1 + i
        for j, val in enumerate(row):
            fmt, _f = fmts[j]
            f = _f or BODY
            al = RIGHT if fmt in (INT_FMT, PCT_FMT, SEC_FMT) else LEFT
            fill = ALT_FILL if i % 2 else None
            _set(ws, rr, left + j, val, font=f, fmt=fmt, fill=fill, align=al)
    return top + 1 + len(rows)  # next free row


def _widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ----------------------------------------------------------------------
def build(analytics_dir):
    analytics_dir = Path(analytics_dir)
    snaps = _read(analytics_dir / "snapshots.csv")
    traffic = _read(analytics_dir / "traffic.csv")
    geo = _read(analytics_dir / "geo.csv")
    search = _read(analytics_dir / "search.csv")
    growth = _read(analytics_dir / "growth.csv")  # ① 주간 스톡 시계열 (별도)
    if not snaps:
        raise RuntimeError("snapshots.csv 없음/빈 파일 — 먼저 `youtube_analytics.py report` 실행")

    dates = sorted({r["measured_on"] for r in snaps})
    latest = dates[-1]
    window = next((r.get("window_days") for r in snaps if r["measured_on"] == latest), "")
    ch = next((r for r in snaps if r["measured_on"] == latest and r["scope"] == "CHANNEL"), {})
    vids = [r for r in snaps if r["measured_on"] == latest and r["scope"] != "CHANNEL"]
    vids.sort(key=lambda r: _num(r.get("views")) or 0, reverse=True)
    tr = [r for r in traffic if r["measured_on"] == latest]
    gl = [r for r in geo if r["measured_on"] == latest]
    sl = [r for r in search if r["measured_on"] == latest]

    wb = Workbook()

    # ================= 1) 개요 =================
    ws = wb.active
    ws.title = "개요"
    ws.sheet_view.showGridLines = False
    _widths(ws, {"A": 22, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
    _set(ws, 1, 1, "Atelier Miku Acappella — 채널 분석 리포트", font=TITLE_F, border=False)
    _set(ws, 2, 1,
         f"측정일 {latest} · 최근 {window}일 · 생성 {dt.date.today().isoformat()} · "
         "출처 YouTube Analytics API",
         font=SUB_F, border=False)
    _set(ws, 3, 1,
         "⚠️ 노출수·CTR은 API에 없음 → 아래 노란 칸에 Studio '도달범위' 값 직접 기입",
         font=SUB_F, border=False)

    # KPI 그리드
    r = 5
    _set(ws, r, 1, "■ 채널 KPI (최근 측정)", font=SEC_F, border=False)
    r += 1
    kpis = [
        ("조회수", _num(ch.get("views")), INT_FMT),
        ("시청 시간(분)", _num(ch.get("watch_minutes")), INT_FMT),
        ("평균 시청(초)", _num(ch.get("avg_view_sec")), SEC_FMT),
        ("평균 시청률", _num(ch.get("avg_view_pct")), PCT_FMT),
        ("좋아요", _num(ch.get("likes")), INT_FMT),
        ("좋아요율", _num(ch.get("like_rate_pct")), PCT_FMT),
        ("댓글", _num(ch.get("comments")), INT_FMT),
        ("구독 기여", _num(ch.get("subs_gained")), INT_FMT),
        ("공유", _num(ch.get("shares")), INT_FMT),
    ]
    col = 1
    for label, val, fmt in kpis:
        _set(ws, r, col, label, font=BOLD, fill=KPI_FILL, align=CENTER)
        _set(ws, r + 1, col, val, font=Font(name=FONT, size=13, bold=True, color=TEAL),
             fmt=fmt, align=CENTER)
        col += 1
        if col > 5:  # 5칸씩 줄바꿈
            col = 1
            r += 2
    r += 3

    # 트래픽 믹스 (요약)
    _set(ws, r, 1, "■ 트래픽 소스 (최근 측정)", font=SEC_F, border=False)
    r += 1
    tr_rows = [[t.get("source_label") or t.get("source_type"),
                _num(t.get("views")), _num(t.get("share_pct")), _num(t.get("avg_view_sec"))]
               for t in tr]
    tr_end = _table(ws, r, 1, ["소스", "조회", "비중", "평균시청"],
                    tr_rows, [(None, None), (INT_FMT, None), (PCT_FMT, None), (SEC_FMT, None)])

    # 지역 / 검색 (오른쪽)
    _set(ws, r, 4, "■ 상위 지역", font=SEC_F, border=False)
    geo_rows = [[g.get("country_label") or g.get("country"),
                 _num(g.get("views")), _num(g.get("avg_view_pct"))] for g in gl]
    geo_end = _table(ws, r + 1, 4, ["지역", "조회", "평균시청률"], geo_rows,
                     [(None, None), (INT_FMT, None), (PCT_FMT, None)])
    sr = max(tr_end, geo_end) + 2

    # Studio 수동 보충 (노란 칸)
    _set(ws, sr, 1, "■ Studio 수동 보충 (API 밖 — 직접 채우기)", font=SEC_F, border=False)
    sr += 1
    for j, h in enumerate(["항목", "값", "메모"]):
        _set(ws, sr, 1 + j, h, font=H_F, fill=H_FILL, align=CENTER if j else LEFT)
    manual = [("노출수 (impressions)", "Studio 도달범위 탭"),
              ("노출 클릭률 CTR (%)", "썸네일 클릭률"),
              ("평균 조회율(%) 노출 대비", "참고")]
    for i, (item, memo) in enumerate(manual):
        rr = sr + 1 + i
        _set(ws, rr, 1, item, font=BODY)
        _set(ws, rr, 2, None, fill=YELLOW)            # ← 사용자 입력 칸
        _set(ws, rr, 3, memo, font=NOTE_F)
    sr = sr + 1 + len(manual) + 2

    # 전략 메모
    _set(ws, sr, 1, "■ 전략 메모 (s340 진단)", font=SEC_F, border=False)
    notes = [
        "쫓지 말 것 = 평균 retention. 구조적임(니치 취향 × 호기심 트래픽 깔때기) · 결함 아님.",
        "레버 ① 고의도 트래픽 더 키우기 (SEO·다국어·보카로 팬덤 타깃).",
        "레버 ② 진성팬 전환 (구독·재생목록·고정댓글). 좋아요율 높음 = fit 강함.",
        "일본 = 최대 미개척. 미쿠 표기로 자연 필터 · 初音ミク/アカペラ 검색 표면 확보.",
    ]
    for i, n in enumerate(notes):
        c = _set(ws, sr + 1 + i, 1, "· " + n, font=BODY, border=False)
        ws.merge_cells(start_row=sr + 1 + i, start_column=1, end_row=sr + 1 + i, end_column=6)
        c.alignment = WRAP

    ws.freeze_panes = "A5"

    # ================= 2) 영상별 =================
    ws2 = wb.create_sheet("영상별")
    ws2.sheet_view.showGridLines = False
    _widths(ws2, {"A": 24, "B": 10, "C": 12, "D": 12, "E": 11, "F": 12,
                  "G": 9, "H": 11, "I": 11})
    _set(ws2, 1, 1, f"영상별 스냅샷 — {latest} (최근 {window}일)", font=TITLE_F, border=False)
    headers = ["영상", "조회", "시청시간(분)", "평균시청", "시청지속률",
               "좋아요", "좋아요율", "댓글", "구독기여"]
    vrows = [[v.get("scope"), _num(v.get("views")), _num(v.get("watch_minutes")),
              _num(v.get("avg_view_sec")), _num(v.get("avg_view_pct")),
              _num(v.get("likes")), _num(v.get("like_rate_pct")),
              _num(v.get("comments")), _num(v.get("subs_gained"))] for v in vids]
    fmts = [(None, None), (INT_FMT, None), (INT_FMT, None), (SEC_FMT, None),
            (PCT_FMT, None), (INT_FMT, None), (PCT_FMT, None), (INT_FMT, None), (INT_FMT, None)]
    end = _table(ws2, 3, 1, headers, vrows, fmts)
    # 합계 행 (SUM 수식 — 안전한 집계)
    _set(ws2, end, 1, "합계", font=BOLD, fill=KPI_FILL)
    for col in (2, 3, 6, 8, 9):  # 조회/시청시간/좋아요/댓글/구독
        L = get_column_letter(col)
        _set(ws2, end, col, f"=SUM({L}4:{L}{end-1})", font=BOLD, fmt=INT_FMT,
             fill=KPI_FILL, align=RIGHT)
    ws2.freeze_panes = "A4"

    # ================= 3) 추이 =================
    ws3 = wb.create_sheet("추이")
    ws3.sheet_view.showGridLines = False
    _set(ws3, 1, 1, "조회수 추이 (측정 회차별)", font=TITLE_F, border=False)
    _set(ws3, 2, 1, "측정을 거듭할수록 행이 쌓이며 추세선이 자라남.", font=SUB_F, border=False)
    scopes = ["CHANNEL"] + [v.get("scope") for v in vids]
    label_of = {"CHANNEL": "채널 전체"}
    pivot = {d: {} for d in dates}
    for row in snaps:
        pivot[row["measured_on"]][row["scope"]] = _num(row.get("views"))
    htop = 4
    _set(ws3, htop, 1, "측정일", font=H_F, fill=H_FILL, align=LEFT)
    for j, sc in enumerate(scopes):
        _set(ws3, htop, 2 + j, label_of.get(sc, sc), font=H_F, fill=H_FILL, align=CENTER)
    for i, d in enumerate(dates):
        rr = htop + 1 + i
        _set(ws3, rr, 1, d, font=BODY, fill=ALT_FILL if i % 2 else None)
        for j, sc in enumerate(scopes):
            _set(ws3, rr, 2 + j, pivot[d].get(sc), font=BODY, fmt=INT_FMT,
                 align=RIGHT, fill=ALT_FILL if i % 2 else None)
    last = htop + len(dates)
    _widths(ws3, {"A": 13})
    for j in range(len(scopes)):
        ws3.column_dimensions[get_column_letter(2 + j)].width = 16

    chart = LineChart()
    chart.title = "조회수 추이"
    chart.style = 12
    chart.y_axis.title = "조회수"
    chart.x_axis.title = "측정일"
    chart.height = 8
    chart.width = 18
    data_ref = Reference(ws3, min_col=2, max_col=1 + len(scopes), min_row=htop, max_row=last)
    cats = Reference(ws3, min_col=1, min_row=htop + 1, max_row=last)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws3.add_chart(chart, f"A{last + 3}")

    # ---- 누적 성장 추이 (lifetime · 채널 성장의 절대 곡선) ----
    # 스톡(odometer)엔 창이 없다 → 두 소스를 측정일 기준 병합해 하나의 촘촘한 곡선으로.
    #  · snapshots.csv CHANNEL 행 = 과거 月1회 report에 얹혀 잡힌 스톡
    #  · growth.csv          = ① 주간 스톡 스냅샷 (`youtube_analytics.py growth`)
    # 같은 날은 growth.csv를 우선(더 정밀한 전용 수집).
    _grow_by_date = {}
    for r in snaps:
        if r.get("scope") == "CHANNEL" and r.get("total_subscribers"):
            _grow_by_date[r["measured_on"]] = r
    for r in growth:
        if r.get("total_subscribers"):
            _grow_by_date[r["measured_on"]] = r
    grow_rows = [_grow_by_date[d] for d in sorted(_grow_by_date)]
    gtop = last + 20  # 조회수 차트 아래
    _set(ws3, gtop, 1, "누적 성장 추이 (lifetime · 측정 회차별)", font=TITLE_F, border=False)
    _set(ws3, gtop + 1, 1,
         "총 구독자·총 조회수·영상수 = 채널 성장의 절대 곡선 (28일 윈도우 활동량과 별개 · 주간 스톡 병합).",
         font=SUB_F, border=False)
    ghtop = gtop + 3
    for j, c in enumerate(["측정일", "총 구독자", "총 조회수", "영상수"]):
        _set(ws3, ghtop, 1 + j, c, font=H_F, fill=H_FILL, align=LEFT if j == 0 else CENTER)
    for i, row in enumerate(grow_rows):
        rr = ghtop + 1 + i
        fill = ALT_FILL if i % 2 else None
        _set(ws3, rr, 1, row["measured_on"], font=BODY, fill=fill)
        _set(ws3, rr, 2, _num(row.get("total_subscribers")), font=BODY, fmt=INT_FMT, align=RIGHT, fill=fill)
        _set(ws3, rr, 3, _num(row.get("total_views")), font=BODY, fmt=INT_FMT, align=RIGHT, fill=fill)
        _set(ws3, rr, 4, _num(row.get("total_videos")), font=BODY, fmt=INT_FMT, align=RIGHT, fill=fill)
    glast = ghtop + len(grow_rows)
    if grow_rows:
        gchart = LineChart()
        gchart.title = "누적 총 구독자 추이"
        gchart.style = 12
        gchart.y_axis.title = "총 구독자"
        gchart.x_axis.title = "측정일"
        gchart.height = 8
        gchart.width = 18
        gdata = Reference(ws3, min_col=2, max_col=2, min_row=ghtop, max_row=glast)
        gcats = Reference(ws3, min_col=1, min_row=ghtop + 1, max_row=glast)
        gchart.add_data(gdata, titles_from_data=True)
        gchart.set_categories(gcats)
        ws3.add_chart(gchart, f"A{glast + 3}")

    # ================= 4) 트래픽 =================
    ws4 = wb.create_sheet("트래픽")
    ws4.sheet_view.showGridLines = False
    _widths(ws4, {"A": 18, "B": 10, "C": 10, "D": 12})
    _set(ws4, 1, 1, f"트래픽 소스 — {latest}", font=TITLE_F, border=False)
    t_end = _table(ws4, 3, 1, ["소스", "조회", "비중", "평균시청"], tr_rows,
                   [(None, None), (INT_FMT, None), (PCT_FMT, None), (SEC_FMT, None)])
    bar = BarChart()
    bar.type = "bar"
    bar.title = "소스별 비중(%)"
    bar.height = 8
    bar.width = 16
    bar.legend = None
    bdata = Reference(ws4, min_col=3, min_row=3, max_row=t_end - 1)
    bcats = Reference(ws4, min_col=1, min_row=4, max_row=t_end - 1)
    bar.add_data(bdata, titles_from_data=True)
    bar.set_categories(bcats)
    ws4.add_chart(bar, f"F3")

    # ================= 5) 지역·검색 =================
    ws5 = wb.create_sheet("지역·검색")
    ws5.sheet_view.showGridLines = False
    _widths(ws5, {"A": 16, "B": 10, "C": 12, "E": 22, "F": 10})
    _set(ws5, 1, 1, f"지역 / 검색어 — {latest}", font=TITLE_F, border=False)
    _table(ws5, 3, 1, ["지역", "조회", "평균시청률"], geo_rows,
           [(None, None), (INT_FMT, None), (PCT_FMT, None)])
    _set(ws5, 3 - 1, 5, "검색 유입 쿼리", font=SEC_F, border=False)
    s_rows = [[s.get("query"), _num(s.get("views"))] for s in sl]
    if not s_rows:
        s_rows = [["(검색 유입 없음)", None]]
    _table(ws5, 3, 5, ["검색어", "조회"], s_rows, [(None, None), (INT_FMT, None)])

    # ================= 6) 원자료 =================
    ws6 = wb.create_sheet("원자료")
    _set(ws6, 1, 1, "snapshots.csv (전체 측정 이력)", font=BOLD, border=False)
    if snaps:
        cols = list(snaps[0].keys())
        for j, h in enumerate(cols):
            _set(ws6, 3, 1 + j, h, font=H_F, fill=H_FILL, align=CENTER)
        for i, row in enumerate(snaps):
            for j, h in enumerate(cols):
                _set(ws6, 4 + i, 1 + j, _num(row.get(h)) if h not in
                     ("measured_on", "start_date", "end_date", "scope", "video_id")
                     else row.get(h), font=BODY)
        for j in range(len(cols)):
            ws6.column_dimensions[get_column_letter(1 + j)].width = 13
        ws6.freeze_panes = "A4"

    out = analytics_dir / "channel_analysis.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    d = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent
    print("생성:", build(d))
